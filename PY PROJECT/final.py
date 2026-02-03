import os
from openpyxl import Workbook, load_workbook
from datetime import datetime
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# all color options
RESET = "\033[0m"
RED = "\033[31m"
GREEN = "\033[32m"
YELLOW = "\033[33m"
BLUE = "\033[34m"
MAGENTA = "\033[35m"
CYAN = "\033[36m"
WHITE = "\033[37m"

def clear_screen():
    os.system('cls' if os.name == 'nt' else 'clear')

def style_excel_sheet(ws):
    header_font = Font(bold=True)
    align = Alignment(vertical="center", wrap_text=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = align
            if cell.row == 1:
                cell.font = header_font
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 4

# from here class management logic start to manage classes
# ===================== CLASS MANAGEMENT =====================
def init_class_management():
    base = "Class_Management"
    removed = os.path.join(base, "removed_classes")
    if not os.path.exists(base):
        os.mkdir(base)
    if not os.path.exists(removed):
        os.mkdir(removed)
    master = os.path.join(base, "class_master.xlsx")
    if not os.path.exists(master):
        wb = Workbook()
        ws = wb.active
        ws.append(["Class_Name", "Total_Sections", "Status"])
        style_excel_sheet(ws)
        wb.save(master)

# -------- ADD NEW CLASS WITH SECTION & SUBJECTS --------
def add_new_class():
    init_class_management()
    base = "Class_Management"
    master_path = os.path.join(base, "class_master.xlsx")
    class_name = input("Enter Class Name (e.g. 8): ").strip()
    section = input("Enter Section (e.g. A): ").upper().strip()
    wb = load_workbook(master_path)
    ws = wb.active
    for r in ws.iter_rows(min_row=2):
        if r[0].value == class_name and r[2].value == "Active":
            print("❌ Class already exists")
            return
    class_folder = os.path.join(base, f"Class_{class_name}")
    os.mkdir(class_folder)
    # subjects
    sub_count = int(input("How many subjects in this class?: "))
    sub_wb = Workbook()
    sub_ws = sub_wb.active
    sub_ws.append(["Subject_ID", "Subject_Name"])
    for i in range(1, sub_count + 1):
        sub = input(f"Enter subject {i}: ")
        sub_ws.append([i, sub])
    style_excel_sheet(sub_ws)
    sub_wb.save(os.path.join(class_folder, "subjects.xlsx"))
    # section file
    sec_wb = Workbook()
    sec_ws = sec_wb.active
    sec_ws.append(["Student_ID", "Name", "DOB", "Age", "Mother's Name","Father's Name", "Address", "Phone No", "Adhar NO"])
    style_excel_sheet(sec_ws)
    sec_wb.save(os.path.join(class_folder, f"section_{section}.xlsx"))
    ws.append([class_name, 1, "Active"])
    wb.save(master_path)
    print(f"✅ Class {class_name}{section} created successfully")

# -------- ADD ONLY SECTION --------
def add_section():
    base = "Class_Management"
    class_name = input("Enter Class Name: ").strip()
    section = input("Enter new Section: ").upper().strip()
    class_folder = os.path.join(base, f"Class_{class_name}")
    if not os.path.exists(class_folder):
        print("❌ Class does not exist")
        return
    sec_path = os.path.join(class_folder, f"section_{section}.xlsx")
    if os.path.exists(sec_path):
        print("❌ Section already exists")
        return
    wb = Workbook()
    ws = wb.active
    ws.append(["Student_ID", "Name", "DOB", "Age", "Mother's Name","Father's Name", "Address", "Phone No", "Adhar NO"])
    wb.save(sec_path)
    master = os.path.join(base, "class_master.xlsx")
    mwb = load_workbook(master)
    mws = mwb.active
    for r in mws.iter_rows(min_row=2):
        if r[0].value == class_name:
            r[1].value += 1
    style_excel_sheet(ws)
    mwb.save(master)
    print(f"✅ Section {section} added to Class {class_name}")

# -------- REMOVE ONLY SECTION --------
def remove_section():
    base = "Class_Management"
    class_name = input("Enter Class Name: ").strip()
    section = input("Enter Section to remove: ").upper().strip()
    sec_path = os.path.join(base, f"Class_{class_name}", f"section_{section}.xlsx")
    if not os.path.exists(sec_path):
        print("❌ Section not found")
        return
    wb = load_workbook(sec_path)
    ws = wb.active
    if ws.max_row > 1:
        print("❌ Section not empty, cannot remove")
        return
    os.remove(sec_path)
    master = os.path.join(base, "class_master.xlsx")
    mwb = load_workbook(master)
    mws = mwb.active
    for r in mws.iter_rows(min_row=2):
        if r[0].value == class_name:
            r[1].value -= 1
    mwb.save(master)
    print(f"✅ Section {section} removed from Class {class_name}")

# -------- REMOVE ENTIRE CLASS --------
def remove_class():
    base = "Class_Management"
    class_name = input("Enter Class Name to remove: ").strip()
    class_folder = os.path.join(base, f"Class_{class_name}")
    if not os.path.exists(class_folder):
        print("❌ Class not found")
        return
    removed_path = os.path.join(base, "removed_classes", f"Class_{class_name}")
    os.rename(class_folder, removed_path)
    master = os.path.join(base, "class_master.xlsx")
    wb = load_workbook(master)
    ws = wb.active
    for r in ws.iter_rows(min_row=2):
        if r[0].value == class_name:
            r[2].value = "Inactive"
    wb.save(master)
    print(f"✅ Class {class_name} removed successfully")

# -------- VIEW CLASSES --------
def view_classes():
    base = "Class_Management"
    master = os.path.join(base, "class_master.xlsx")
    wb = load_workbook(master)
    ws = wb.active
    print("\n📚 CLASSES LIST\n")
    for r in ws.iter_rows(min_row=2, values_only=True):
        print(f"Class: {r[0]} | Sections: {r[1]} | Status: {r[2]}")

# -------- CLASS MANAGEMENT MENU --------
def classmgmnt():
    init_class_management()
    while True:
        clear_screen()
        print("\n===== CLASS MANAGEMENT =====")
        print("1. Add New Class with Section & Subjects")
        print("2. Add Section to Existing Class")
        print("3. Remove Entire Class")
        print("4. Remove Only Section")
        print("5. View Classes")
        print("6. Back to Main Menu")
        choice = input("Enter choice (1-6): ")
        if choice == '1':
            add_new_class()
        elif choice == '2':
            add_section()
        elif choice == '3':
            remove_class()
        elif choice == '4':
            remove_section()
        elif choice == '5':
            view_classes()
        elif choice == '6':
            admin()
        else:
            print("❌ Invalid choice")


#  Student management logics are here all the add update delete view logics are here
# this function is for calculating age with current time if user give input to its date of birth
def calculate_age(dob):
    dob = datetime.strptime(dob, "%Y-%m-%d")
    today = datetime.today()
    age = today.year - dob.year
    if (today.month, today.day) < (dob.month, dob.day):
        age -= 1
    return age

# this function is for adding student single or multiple
def add_student():
    classname = input("Enter class name: ").strip()
    section = input("Enter section: ").upper().strip()
    section_file = os.path.join(
        "Class_Management", f"Class_{classname}", f"section_{section}.xlsx"
    )
    if not os.path.exists(section_file):
        print("❌ Class or Section does not exist")
        return
    wb = load_workbook(section_file)
    ws = wb.active
    print("\n1. Add Single Student")
    print("2. Add Multiple Students")
    mode = input("Enter choice (1/2): ").strip()
    if mode == '1':
        count = 1
    elif mode == '2':
        count = int(input("How many students do you want to add?: "))
    else:
        print("❌ Invalid choice")
        return
    for _ in range(count):
        sid = ws.max_row
        print(f"\n--- Student ID: {sid} ---")
        name = input("Name: ")
        dob = input("DOB (YYYY-MM-DD): ")
        age = calculate_age(dob)
        mname = input("Mother Name: ")
        fname = input("Father Name: ")
        address = input("Address: ")
        phone = input("Phone: ")
        adhar = input("Aadhar: ")
        ws.append([
            sid, name, dob, age, mname, fname, address, phone, adhar
        ])
    wb.save(section_file)
    print("✅ Student(s) added successfully")

# To remove student data from excel sheet
def remove_student():
    print("You have selected to remove a student.")
    classname = input("Enter class name : ").strip()
    section = input("Enter section : ").upper().strip()
    sid = int(input("Enter student ID to remove : "))
    section_file = os.path.join(
        "Class_Management",
        f"Class_{classname}",
        f"section_{section}.xlsx"
    )
    if not os.path.exists(section_file):
        print("❌ Class or Section does not exist")
        return
    wb = load_workbook(section_file)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[0].value == sid:
            ws.delete_rows(row[0].row)
            wb.save(section_file)
            print("✅ Student removed successfully")
            return
    print("❌ Student ID not found")

# functions to view students at different stages
# this function helps to view single student details
def single_view():
    classname = input("Enter class name : ").strip()
    section = input("Enter section : ").upper().strip()
    sid = int(input("Enter Student ID : "))
    section_file = os.path.join(
        "Class_Management",
        f"Class_{classname}",
        f"section_{section}.xlsx"
    )
    wb = load_workbook(section_file)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == sid:
            print(f"""
ID      : {row[0]}
Name    : {row[1]}
DOB     : {row[2]}
Age     : {row[3]}
Mother  : {row[4]}
Father  : {row[5]}
Address : {row[6]}
Phone   : {row[7]}
Aadhar  : {row[8]}
-------------------------
""")
            return
    print("❌ Student not found")

# to see complete class students
def mulit_view():
    classname = input("Enter class name : ").strip()
    section = input("Enter section : ").upper().strip()
    section_file = os.path.join(
        "Class_Management",
        f"Class_{classname}",
        f"section_{section}.xlsx"
    )
    wb = load_workbook(section_file)
    ws = wb.active
    print(f"\n📋 Students of Class {classname}{section}\n")
    for row in ws.iter_rows(min_row=2, values_only=True):
        print(f"""
ID      : {row[0]}
Name    : {row[1]}
DOB     : {row[2]}
Age     : {row[3]}
Mother  : {row[4]}
Father  : {row[5]}
Address : {row[6]}
Phone   : {row[7]}
Aadhar  : {row[8]}
-------------------------
""")

# function that operate all these views singleview and multiview
def view_students():
    print("You have selected to view all students.")
    print("Want to see single student details or all students details?")
    print("1. Single Student")
    print("2. All Students")
    choice = input("Enter your choice (1 or 2): ")
    if choice == '1':
        single_view()
    elif choice == '2':
        mulit_view()
    else :
        print("Invalid Selection")

# for updating student information
def update_student_info():
    classname = input("Enter class name : ").strip()
    section = input("Enter section : ").upper().strip()
    sid = int(input("Enter student ID : "))
    section_file = os.path.join(
        "Class_Management",
        f"Class_{classname}",
        f"section_{section}.xlsx"
    )
    wb = load_workbook(section_file)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[0].value == sid:
            print("""
            Select what you want to update:
                1. Name
                2. DOB
                3. Address
                4. Phone
            """)
            choice = input("Enter choice : ")
            if choice == '1':
                row[1].value = input("Enter new name: ")
            elif choice == '2':
                dob = input("Enter new DOB (YYYY-MM-DD): ")
                row[2].value = dob
                row[3].value = calculate_age(dob)
            elif choice == '3':
                row[6].value = input("Enter new address: ")
            elif choice == '4':
                row[7].value = input("Enter new phone: ")
            else:
                print("❌ Invalid choice")
                return
            wb.save(section_file)
            print("✅ Student information updated successfully")
            return
    print("❌ Student ID not found")

# this manages all the student functions like add remove and update info with menu driven
def stdntmgmnt():
    while True:
        clear_screen()
        print("You have selected Student Management")
        print("Now select which operation you want to perform")
        print("1. Add Student")
        print("2. Remove Student")
        print("3. View Students")
        print("4. Update Student Information")
        print("5. Back to Main Menu")
        choice = (input("Enter your choice (1-6): "))
        if choice == '1':
            add_student()
        elif choice == '2':
            remove_student()
        elif choice == '3':
            view_students()
        elif choice == '4':
            update_student_info()
        elif choice == '5':
            admin()
        else:
            print("Invalid choice. Please select a number between 1 and 6.")

# Teacher management logics are here all the add update delete view logics are here
# this function generates automatic id
def generate_teacher_id():
    folder = "Teacher_Management"
    id_file = os.path.join(folder, "teacher_id.txt")
    if not os.path.exists(id_file):
        with open(id_file, "w") as f:
            f.write("1000")
    with open(id_file, "r") as f:
        last_id = int(f.read())
    new_id = last_id + 1
    with open(id_file, "w") as f:
        f.write(str(new_id))
    return new_id

# For initializing excel sheet as it automatically generates the excel sheet when admin select teacher management
def init_teacher_files():
    folder = "Teacher_Management"
    if not os.path.exists(folder):
        os.mkdir(folder)
    id_file = os.path.join(folder, "teacher_id.txt")
    if not os.path.exists(id_file):
        with open(id_file, "w") as f:
            f.write("1000")
    files = {
        "teacher_details.xlsx": [
        "ID", "Name","Password", "DOB", "Age", "Phone", "Email",
        "Address", "Aadhar", "Account No", "Bank"],
        "teacher_salary.xlsx": [
        "ID", "Name", "Month",
        "Basic", "Incentive", "Gross",
        "PF", "Professional_Tax", "TDS",
        "Total_Deduction", "Net_Salary"],
        "removed_teachers.xlsx": [
        "ID", "Name", "Reason", "Removed Date"]
    }
    for file, headers in files.items():
        path = os.path.join(folder, file)
        if not os.path.exists(path):
            wb = Workbook()
            ws = wb.active
            ws.append(headers)
            style_excel_sheet(ws)
            wb.save(path)

# this if for adding teachers data
def add_teacher():
    folder = "Teacher_Management"
    path = os.path.join(folder, "teacher_details.xlsx")
    wb = load_workbook(path)
    ws = wb.active
    tid = generate_teacher_id()
    name = input("Enter Name: ")
    password=input("Enter Password")
    dob = input("Enter Date of birth (YYYY-MM-DD): ")
    age = calculate_age(dob)
    phno = input("Enter Phone Number: ")
    email = input("Enter Email: ")
    address = input("Enter the address: ")
    adhar = input("Enter Aadhar number: ")
    accountnumber = input("Enter Account Number: ")
    nameofbank = input("Enter Bank Name: ")
    ws.append([tid, name,password, dob, age, phno, email, address, adhar, accountnumber, nameofbank])
    wb.save(path)
    print(f"✅ Teacher added successfully with ID {tid}")

# this is for removing teacher and it stores in seperate excel sheet
def remove_teacher():
    from datetime import datetime
    folder = "Teacher_Management"
    main_file = os.path.join(folder, "teacher_details.xlsx")
    removed_file = os.path.join(folder, "removed_teachers.xlsx")
    tid = int(input("Enter Teacher ID: "))
    reason = input("Reason for removing: ")
    wb = load_workbook(main_file)
    ws = wb.active
    rwb = load_workbook(removed_file)
    rws = rwb.active
    # safety check – already removed?
    for r in rws.iter_rows(min_row=2):
        if r[0].value == tid:
            print("⚠️ Teacher already removed earlier")
            return
    for row in ws.iter_rows(min_row=2):
        if row[0].value == tid:
            rws.append([
                row[0].value,
                row[1].value,
                reason,
                datetime.today().strftime("%Y-%m-%d")
            ])
            ws.delete_rows(row[0].row)
            wb.save(main_file)
            rwb.save(removed_file)
            print("✅ Teacher removed successfully")
            return
    print("❌ Teacher ID not found")

# functions for viewing
# sirf ek specific class ke teachers hi dekh paoge/
# to view teachers of specific class
def clsview():
    cls = input("Enter the class you want to view teachers for: ").strip()
    path = os.path.join("Teacher_Management", "teacher_assignments.xlsx")
    wb = load_workbook(path)
    ws = wb.active
    found = False
    print(f"\n📚 Teachers for Class {cls}\n")
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2] == cls:
            found = True
            print(f"""
                ID      : {row[0]}
                Name    : {row[1]}
                Section : {row[3]}
                Subject : {row[4]}
                Role    : {row[5]}
                -------------------------
""")
    if not found:
        print("⚠️ No teachers found for this class")

# to view all subject teachers
def subjectview():
    subject = input("Enter the subject you want to view teachers for: ").strip().lower()
    path = os.path.join("Teacher_Management", "teacher_assignments.xlsx")
    wb = load_workbook(path)
    ws = wb.active
    found = False
    print(f"\n📘 Teachers for Subject: {subject.title()}\n")
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[4].lower() == subject:
            found = True
            print(f"""
                ID      : {row[0]}
                Name    : {row[1]}
                Class   : {row[2]}
                Section : {row[3]}
                Role    : {row[5]}
                -------------------------
""")
    if not found:
        print("⚠️ No teachers found for this subject")

# to view all teachers
def allview():
    folder = "Teacher_Management"
    path = os.path.join(folder, "teacher_details.xlsx")
    wb = load_workbook(path)
    ws = wb.active
    if ws.max_row == 1:
        print("⚠️ No teachers found")
        return
    print("\n📋 ALL TEACHERS LIST\n")
    for row in ws.iter_rows(min_row=2, values_only=True):
        print(f"""
                ID      : {row[0]}
                Name    : {row[1]}
                DOB     : {row[3]}
                Age     : {row[4]}
                Phone   : {row[5]}
                Email   : {row[6]}
                Address : {row[7]}
                Aadhar  : {row[8]}
                Bank    : {row[10]}
                -------------------------
            """)

# to see only class teachers
def clsteacherview():
    path = os.path.join("Teacher_Management", "teacher_assignments.xlsx")
    wb = load_workbook(path)
    ws = wb.active
    found = False
    print("\n🏫 CLASS TEACHERS LIST\n")
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[5] == "Class Teacher":
            found = True
            print(f"""
            Name    : {row[1]}
            ID      : {row[0]}
            Class   : {row[2]}
            Section : {row[3]}
            -------------------------
            """)
    if not found:
        print("⚠️ No class teachers assigned yet")

# manages all views
def view_teachers():
    print("Select from the following")
    print("1. According to class wise ")
    print("2. According to subject wise ")
    print("3. All Class teachers")
    print("4. All Teachers")
    print("5. Return Back")
    choice = input("Enter your choice ")
    if choice == '1':
        clsview()
    elif choice == '2':
        subjectview()
    elif choice =='3':
        clsteacherview()
    elif choice =='4':
        allview()
    elif choice =='5':
        teachermgmnt()
    else:
        print("Invalid Output")

# this function initializes new file that store classes that are assignes to teachers
def init_teacher_assignment_file():
    folder = "Teacher_Management"
    path = os.path.join(folder, "teacher_assignments.xlsx")
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.append([
            "Teacher_ID",
            "Teacher_Name",
            "Class",
            "Section",
            "Subject",
            "Role"
        ])
        style_excel_sheet(ws)
        wb.save(path)

# and this function is to assign classes to teachers
def assign_classes_to_teachers():
    init_teacher_assignment_file()
    tid = int(input("Enter Teacher ID: "))
    class_name = input("Enter Class: ").strip()
    section = input("Enter Section: ").upper().strip()
    subject = input("Enter Subject: ").strip()
    role_choice = input("Assign as (1) Class Teacher or (2) Subject Teacher: ")
    # ---------- VERIFY TEACHER ----------
    t_path = os.path.join("Teacher_Management", "teacher_details.xlsx")
    twb = load_workbook(t_path)
    tws = twb.active
    teacher_name = None
    for row in tws.iter_rows(min_row=2):
        if row[0].value == tid:
            teacher_name = row[1].value
            break
    if not teacher_name:
        print("❌ Teacher not found")
        return
    # ---------- VERIFY CLASS & SECTION ----------
    class_folder = os.path.join("Class_Management", f"Class_{class_name}")
    section_file = os.path.join(class_folder, f"section_{section}.xlsx")
    if not os.path.exists(class_folder):
        print("❌ Class does not exist")
        return
    if not os.path.exists(section_file):
        print("❌ Section does not exist")
        return
    # ---------- VERIFY SUBJECT ----------
    subject_file = os.path.join(class_folder, "subjects.xlsx")
    swb = load_workbook(subject_file)
    sws = swb.active
    subject_found = False
    for row in sws.iter_rows(min_row=2):
        if row[1].value.lower() == subject.lower():
            subject_found = True
            break
    if not subject_found:
        print("❌ Subject not found in this class")
        return
    # ---------- LOAD ASSIGNMENT FILE ----------
    a_path = os.path.join("Teacher_Management", "teacher_assignments.xlsx")
    awb = load_workbook(a_path)
    aws = awb.active
    # ---------- CLASS TEACHER RULES ----------
    if role_choice == '1':
        # Rule 1: Class already has class teacher?
        for r in aws.iter_rows(min_row=2):
            if (
                r[2].value == class_name and
                r[3].value == section and
                r[5].value == "Class Teacher"
            ):
                print("❌ This class already has a class teacher")
                return
        # Rule 2: Teacher already class teacher somewhere else?
        for r in aws.iter_rows(min_row=2):
            if (
                r[0].value == tid and
                r[5].value == "Class Teacher"
            ):
                print("❌ Teacher is already a class teacher of another class")
                return
        role = "Class Teacher"
    elif role_choice == '2':
        role = "Subject Teacher"
    else:
        print("❌ Invalid role selection")
        return
    # ---------- SAVE ASSIGNMENT ----------
    aws.append([
        tid,
        teacher_name,
        class_name,
        section,
        subject,
        role
    ])
    awb.save(a_path)
    print("🎉 Teacher assignment completed successfully")

# to update information of teachers
def update_teacher_info():
    folder = "Teacher_Management"
    path = os.path.join(folder, "teacher_details.xlsx")
    tid = int(input("Enter the ID of teacher that you want to update: "))
    wb = load_workbook(path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[0].value == tid:
            print("""
                    Select what you want to update:
                    1. Name
                    2. Date of Birth
                    3. Phone Number
                    5. Address
                    4. Email
                    6. Bank Account Number
                    7. Bank Name
                """)
            choice = input("Enter your choice (1-7): ")
            if choice == '1':
                row[1].value = input("Enter new Name: ")
            elif choice == '2':
                dob = input("Enter new DOB (YYYY-MM-DD): ")
                row[2].value = dob
                row[3].value = calculate_age(dob)
            elif choice == '3':
                row[4].value = input("Enter new Phone Number: ")
            elif choice == '4':
                row[5].value = input("Enter new Email: ")
            elif choice == '5':
                row[6].value = input("Enter new Address: ")
            elif choice == '6':
                row[8].value = input("Enter new Account Number: ")
            elif choice == '7':
                row[9].value = input("Enter new Bank Name: ")
            else:
                print("❌ Invalid selection")
                return
            wb.save(path)
            print("✅ Teacher information updated successfully")
            return
    print("❌ Teacher ID not found")

# managing salary of teachers
def manage_teacher_salary():
    folder = "Teacher_Management"
    salary_path = os.path.join(folder, "teacher_salary.xlsx")
    assign_path = os.path.join(folder, "teacher_assignments.xlsx")
    teacher_path = os.path.join(folder, "teacher_details.xlsx")
    tid = int(input("Enter Teacher ID: "))
    month = input("Enter Salary Month (e.g. July-2026): ")
    # ---------- GET TEACHER NAME ----------
    twb = load_workbook(teacher_path)
    tws = twb.active
    teacher_name = None
    for r in tws.iter_rows(min_row=2):
        if r[0].value == tid:
            teacher_name = r[1].value
            break
    if not teacher_name:
        print("❌ Teacher not found")
        return
    # ---------- CHECK ROLE ----------
    awb = load_workbook(assign_path)
    aws = awb.active
    is_class_teacher = False
    for r in aws.iter_rows(min_row=2):
        if r[0].value == tid and r[5].value == "Class Teacher":
            is_class_teacher = True
            break
    # ---------- SALARY STRUCTURE ----------
    incentive = 10000
    if is_class_teacher:
        basic = 55000
    else:
        basic = 40000
    gross = basic + incentive
    pf = basic * 0.12
    pt = 200
    tds = gross * 0.05
    total_deduction = pf + pt + tds
    net_salary = gross - total_deduction
    # ---------- STORE IN EXCEL ----------
    swb = load_workbook(salary_path)
    sws = swb.active
    sws.append([
        tid, teacher_name, month,
        basic, incentive, gross,
        pf, pt, tds,
        total_deduction, net_salary
    ])
    swb.save(salary_path)
    print("✅ Salary calculated and stored successfully")
    print(f"💰 Net Salary: ₹{net_salary}")

# generate salary slips in pdf format
def generate_salary_slip():
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    )
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    folder = "Teacher_Management"
    salary_path = os.path.join(folder, "teacher_salary.xlsx")
    tid = int(input("Enter Teacher ID: "))
    month = input("Enter Month (e.g. July-2026): ")
    wb = load_workbook(salary_path)
    ws = wb.active
    record = None
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] == tid and r[2] == month:
            record = r
            break
    if not record:
        print("❌ Salary record not found")
        return
    pdf_name = f"SalarySlip_{record[1]}_{month}.pdf"
    pdf_path = os.path.join(folder, pdf_name)
    doc = SimpleDocTemplate(
        pdf_path,
        rightMargin=36, leftMargin=36,
        topMargin=36, bottomMargin=36
    )
    styles = getSampleStyleSheet()
    elements = []
    # ---------- TITLE ----------
    elements.append(Paragraph("<b>SALARY SLIP</b>", styles["Title"]))
    elements.append(Spacer(1, 15))
    # ---------- BASIC INFO ----------
    info_table = Table([
        ["Teacher Name", record[1]],
        ["Teacher ID", record[0]],
        ["Salary Month", record[2]]
    ], colWidths=[150, 250])
    info_table.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('BACKGROUND', (0,0), (0,-1), colors.lightgrey),
        ('FONT', (0,0), (0,-1), 'Helvetica-Bold'),
        ('PADDING', (0,0), (-1,-1), 8)
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 20))
    # ---------- SALARY DETAILS ----------
    salary_table = Table([
        ["Basic Salary", f"₹{record[3]}"],
        ["Incentive", f"₹{record[4]}"],
        ["Gross Salary", f"₹{record[5]}"],
        ["PF (12%)", f"₹{record[6]}"],
        ["Professional Tax", f"₹{record[7]}"],
        ["TDS (5%)", f"₹{record[8]}"],
        ["Total Deduction", f"₹{record[9]}"],
        ["NET SALARY", f"₹{record[10]}"]
    ], colWidths=[200, 200])
    salary_table.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('BACKGROUND', (0,0), (0,-2), colors.whitesmoke),
        ('BACKGROUND', (0,-1), (-1,-1), colors.lightgrey),
        ('FONT', (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('ALIGN', (1,0), (-1,-1), 'RIGHT'),
        ('PADDING', (0,0), (-1,-1), 8)
    ]))
    elements.append(salary_table)
    doc.build(elements)
    print(f"✅ Salary slip generated: {pdf_name}")

# for managing teachers data by admin
def teachermgmnt():
    init_teacher_files()
    while True:
        clear_screen()
        print("You have selected Teacher Management")
        print("Now select which operation you want to perform")
        print("1. Add Teacher")
        print("2. Remove Teacher")
        print("3. View Teachers")
        print("4. Assign Classes to Teachers")
        print("5. Update Teacher Information")
        print("6. Salary Management")
        print("7. Generate Salary Slip")
        print("8. Back to Main Menu")
        choice = (input("Enter your choice (1-8): "))
        if choice == '1':
            add_teacher()
        elif choice == '2':
            remove_teacher()
        elif choice == '3':
            view_teachers()
        elif choice == '4':
            assign_classes_to_teachers()
        elif choice == '5':
            update_teacher_info()
        elif choice == '6':
            manage_teacher_salary()
        elif choice == '7':
            generate_salary_slip()
        elif choice == '8':
            admin()
        else:
            print("Invalid choice. Please select a number between 1 and 8.")

# New module for teacher operations
# Helper function to login in to teachers
def teacher_login(username, password):
    path = os.path.join("Teacher_Management", "teacher_details.xlsx")
    if not os.path.exists(path):
        print("❌ Teacher details file not found")
        return None
    wb = load_workbook(path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[1]      # Column B
        pwd  = row[2]      # Column C
        if str(name).strip().lower() == username.strip().lower() \
            and str(pwd) == str(password):
            return name     # ✅ RETURN TEACHER NAME
    return None

# to see assigned classes of teacher
def view_assigned_classes(teacher_name):
    path = os.path.join("Teacher_Management", "teacher_assignments.xlsx")
    if not os.path.exists(path):
        print("❌ No assignments found")
        return
    wb = load_workbook(path)
    ws = wb.active
    found = False
    print(f"\n📚 Assigned Classes for {teacher_name}\n")
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1].strip().lower() == teacher_name.strip().lower():
            found = True
            print(f"""
Class   : {row[2]}
Section : {row[3]}
Subject : {row[4]}
Role    : {row[5]}
-------------------------
""")
    if not found:
        print("⚠️ No classes assigned yet")


# to update informations like phone no address
def update_teacher_profile(teacher_name):
    path = os.path.join("Teacher_Management", "teacher_details.xlsx")
    wb = load_workbook(path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[1].value.strip().lower() == teacher_name.strip().lower():
            print("""
Select what you want to update:
1. Phone Number
2. Email
3. Address
4. Password
""")
            choice = input("Enter choice (1-4): ")
            if choice == '1':
                row[5].value = input("Enter new phone number: ")
            elif choice == '2':
                row[6].value = input("Enter new email: ")
            elif choice == '3':
                row[7].value = input("Enter new address: ")
            elif choice == '4':
                row[2].value = input("Enter new password: ")
            else:
                print("❌ Invalid choice")
                return
            wb.save(path)
            print("✅ Profile updated successfully")
            return
    print("❌ Teacher record not found")

# to ensure that attendence file or folder is ready or not
def ensure_attendance_file(class_name, section):
    base = "Attendance"
    class_folder = os.path.join(base, f"Class_{class_name}")
    if not os.path.exists(base):
        os.mkdir(base)
    if not os.path.exists(class_folder):
        os.mkdir(class_folder)
    att_file = os.path.join(class_folder, f"section_{section}.xlsx")
    if not os.path.exists(att_file):
        student_file = os.path.join(
            "Class_Management",
            f"Class_{class_name}",
            f"section_{section}.xlsx"
        )
        if not os.path.exists(student_file):
            print("❌ Class/Section not found in Student records")
            return None
        swb = load_workbook(student_file)
        sws = swb.active
        awb = Workbook()
        aws = awb.active
        aws.append(["Student_ID", "Name"])
        for row in sws.iter_rows(min_row=2, values_only=True):
            aws.append([row[0], row[1]])
        style_excel_sheet(aws)
        awb.save(att_file)
    return att_file

# for marking attendence
def mark_attendance():
    class_name = input("Enter Class: ").strip()
    section = input("Enter Section: ").upper().strip()
    today = datetime.today().strftime("%Y-%m-%d")
    att_file = ensure_attendance_file(class_name, section)
    if not att_file:
        return
    wb = load_workbook(att_file)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    if today not in headers:
        ws.cell(row=1, column=len(headers) + 1).value = today
    date_col = headers.index(today) + 1 if today in headers else ws.max_column
    print(f"\n📝 Marking Attendance for Class {class_name}{section} ({today})\n")
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row, 2).value
        status = input(f"{name} (P/A): ").strip().upper()
        while status not in ['P', 'A']:
            status = input("❌ Enter only P or A: ").strip().upper()
        ws.cell(row=row, column=date_col).value = status
    wb.save(att_file)
    print("✅ Attendance recorded successfully")

# for managing attendence function
def attendance_management():
    while True:
        clear_screen()
        print("\n===== ATTENDANCE MANAGEMENT =====")
        print("1. Mark Attendance")
        print("2. Back")
        choice = input("Enter choice (1-2): ")
        if choice == '1':
            mark_attendance()
        elif choice == '2':
            break
        else:
            print("❌ Invalid choice")

# Grading on basis of marks that you receive
def calculate_grade(percent):
    if percent >= 90:
        return "A1"
    elif percent >= 80:
        return "A2"
    elif percent >= 70:
        return "B1"
    elif percent >= 60:
        return "B2"
    elif percent >= 50:
        return "C"
    elif percent >= 40:
        return "D"
    else:
        return "E"

# to see marks file and folder are ready or not
def ensure_marks_file(class_name, section):
    base = "Marks"
    class_folder = os.path.join(base, f"Class_{class_name}")
    if not os.path.exists(base):
        os.mkdir(base)
    if not os.path.exists(class_folder):
        os.mkdir(class_folder)
    marks_file = os.path.join(class_folder, f"section_{section}.xlsx")
    if not os.path.exists(marks_file):
        student_file = os.path.join(
            "Class_Management",
            f"Class_{class_name}",
            f"section_{section}.xlsx"
        )
        if not os.path.exists(student_file):
            print("❌ Class/Section not found")
            return None
        swb = load_workbook(student_file)
        sws = swb.active
        mwb = Workbook()
        mws = mwb.active
        mws.append(["Student_ID", "Name"])
        for row in sws.iter_rows(min_row=2, values_only=True):
            mws.append([row[0], row[1]])
        style_excel_sheet(mws)
        mwb.save(marks_file)
    return marks_file

# to create exams
def create_exam():
    class_name = input("Enter Class: ").strip()
    section = input("Enter Section: ").upper().strip()
    exam_name = input("Exam Name: ").strip()
    start_date = input("Exam Start Date (YYYY-MM-DD): ").strip()
    max_marks = int(input("Max marks per subject: "))
    marks_file = ensure_marks_file(class_name, section)
    wb = load_workbook(marks_file)
    ws = wb.active
    # get subjects
    sub_file = os.path.join(
        "Class_Management", f"Class_{class_name}", "subjects.xlsx"
    )
    swb = load_workbook(sub_file)
    sws = swb.active
    subjects = [r[1] for r in sws.iter_rows(min_row=2, values_only=True)]
    start_row = ws.max_row + 2
    ws.merge_cells(start_row=start_row, start_column=1,
                end_row=start_row, end_column=len(subjects)+5)
    ws.cell(start_row, 1).value = (
        f"Exam: {exam_name} | Start Date: {start_date} | Max Marks: {max_marks}"
    )
    headers = ["Student_ID", "Name"] + subjects + ["Total", "Percentage", "Grade"]
    ws.append(headers)
    student_file = os.path.join(
        "Class_Management", f"Class_{class_name}", f"section_{section}.xlsx"
    )
    swb2 = load_workbook(student_file)
    sws2 = swb2.active
    for s in sws2.iter_rows(min_row=2, values_only=True):
        ws.append([s[0], s[1]] + [""] * (len(subjects) + 3))
    wb.save(marks_file)
    print("✅ Exam created successfully")

# to entering marks in created exams
def enter_marks():
    class_name = input("Enter Class: ").strip()
    section = input("Enter Section: ").upper().strip()
    exam_name = input("Exam Name: ").strip()
    marks_file = ensure_marks_file(class_name, section)
    wb = load_workbook(marks_file)
    ws = wb.active
    # 🔍 Locate exam block
    exam_row = None
    for r in range(1, ws.max_row + 1):
        val = ws.cell(r, 1).value
        if isinstance(val, str) and val.startswith("Exam:") and exam_name in val:
            exam_row = r
            break
    if not exam_row:
        print("❌ Exam not found")
        return
    header_row = exam_row + 1
    headers = [c.value for c in ws[header_row]]
    subjects = headers[2:-3]
    max_marks = int(ws.cell(exam_row, 1).value.split("Max Marks: ")[1])
    print("\n1. Enter marks for Single Student")
    print("2. Enter marks for Whole Class")
    mode = input("Enter choice (1/2): ").strip()
    if mode == '1':
        sid = int(input("Enter Student ID: "))
        rows = [r for r in range(header_row + 1, ws.max_row + 1)
                if ws.cell(r, 1).value == sid]
        if not rows:
            print("❌ Student not found")
            return
    elif mode == '2':
        rows = range(header_row + 1, ws.max_row + 1)
    else:
        print("❌ Invalid choice")
        return
    for r in rows:
        if ws.cell(r, 1).value is None:
            continue
        print(f"\nEntering marks for {ws.cell(r,2).value}")
        total = 0
        for i, subject in enumerate(subjects):
            col = 3 + i
            while True:
                marks = input(f"{subject} (0–{max_marks}): ")
                if marks.isdigit() and 0 <= int(marks) <= max_marks:
                    marks = int(marks)
                    break
                print("❌ Invalid marks")
            ws.cell(r, col).value = marks
            total += marks
        percent = (total / (max_marks * len(subjects))) * 100
        ws.cell(r, col + 1).value = total
        ws.cell(r, col + 2).value = round(percent, 2)
        ws.cell(r, col + 3).value = calculate_grade(percent)
    wb.save(marks_file)
    print("✅ Marks entered successfully")

# to update the given marks if they are given wrong
def update_marks():
    class_name = input("Enter Class: ").strip()
    section = input("Enter Section: ").upper().strip()
    exam_name = input("Exam Name: ").strip()
    sid = int(input("Enter Student ID: "))
    marks_file = ensure_marks_file(class_name, section)
    wb = load_workbook(marks_file)
    ws = wb.active
    exam_row = None
    # 🔍 Locate exam block
    for r in range(1, ws.max_row + 1):
        val = ws.cell(r, 1).value
        if isinstance(val, str) and val.startswith("Exam:") and exam_name in val:
            exam_row = r
            break
    if not exam_row:
        print("❌ Exam not found")
        return
    header_row = exam_row + 1
    headers = [c.value for c in ws[header_row]]
    subjects = headers[2:-3]  # subject columns
    max_marks = int(ws.cell(exam_row, 1).value.split("Max Marks: ")[1])
    # 🔍 Find student row
    student_row = None
    for r in range(header_row + 1, ws.max_row + 1):
        if ws.cell(r, 1).value == sid:
            student_row = r
            break
    if not student_row:
        print("❌ Student not found")
        return
    print("\nSelect subject to update:")
    for i, sub in enumerate(subjects, start=1):
        print(f"{i}. {sub}")
    choice = int(input("Enter choice: "))
    col = 2 + choice
    new_marks = int(input(f"Enter new marks for {subjects[choice-1]}: "))
    ws.cell(student_row, col).value = new_marks
    # 🔄 Recalculate total, percent, grade
    total = sum(ws.cell(student_row, c).value or 0
                for c in range(3, 3 + len(subjects)))
    percent = (total / (max_marks * len(subjects))) * 100
    ws.cell(student_row, 3 + len(subjects)).value = total
    ws.cell(student_row, 4 + len(subjects)).value = round(percent, 2)
    ws.cell(student_row, 5 + len(subjects)).value = calculate_grade(percent)
    wb.save(marks_file)
    print("✅ Marks updated successfully")

# to generate report card of all exams
def generate_report_card_pdf():
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib import colors
    class_name = input("Enter Class: ").strip()
    section = input("Enter Section: ").upper().strip()
    sid = int(input("Enter Student ID: "))
    marks_file = os.path.join(
        "Marks", f"Class_{class_name}", f"section_{section}.xlsx"
    )
    if not os.path.exists(marks_file):
        print("❌ Marks file not found")
        return
    wb = load_workbook(marks_file)
    ws = wb.active
    base_folder = "Report_Cards"
    class_folder = os.path.join(base_folder, f"Class_{class_name}")
    section_folder = os.path.join(class_folder, f"Section_{section}")
    os.makedirs(section_folder, exist_ok=True)
    pdf_path = os.path.join(
        section_folder, f"ReportCard_{class_name}{section}_{sid}.pdf"
    )
    # ---------- DOCUMENT SETUP ----------
    doc = SimpleDocTemplate(
        pdf_path,
        rightMargin=36, leftMargin=36,
        topMargin=36, bottomMargin=36
    )
    styles = getSampleStyleSheet()
    elements = []
    # ---------- TITLE ----------
    title_style = ParagraphStyle(
        "TitleCenter",
        parent=styles["Title"],
        alignment=TA_CENTER
    )
    elements.append(Paragraph("REPORT CARD", title_style))
    elements.append(Spacer(1, 14))
    # ---------- STUDENT INFO BOX ----------
    info_table = Table(
        [
            ["Class", class_name],
            ["Section", section],
            ["Student ID", sid],
        ],
        colWidths=[120, 200]
    )
    info_table.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('BACKGROUND', (0,0), (0,-1), colors.lightgrey),
        ('FONT', (0,0), (0,-1), 'Helvetica-Bold'),
        ('PADDING', (0,0), (-1,-1), 8),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 20))
    # ---------- PROCESS EXAMS ----------
    combined_total = 0
    combined_max = 0
    final_fail = False
    found_student = False
    r = 1
    while r <= ws.max_row:
        exam_header = ws.cell(r, 1).value
        if isinstance(exam_header, str) and exam_header.startswith("Exam:"):
            header_row = r + 1
            headers = [c.value for c in ws[header_row]]
            max_marks = int(exam_header.split("Max Marks: ")[1])
            subject_count = len(headers) - 5
            exam_max_total = max_marks * subject_count
            for s in range(header_row + 1, ws.max_row + 1):
                if ws.cell(s, 1).value == sid:
                    found_student = True
                    row_values = [
                        ws.cell(s, c + 1).value
                        for c in range(len(headers))
                    ]
                    total = row_values[-3]
                    percent = row_values[-2]
                    grade = row_values[-1]
                    combined_total += total
                    combined_max += exam_max_total
                    if grade == "E":
                        final_fail = True
                    elements.append(
                        Paragraph(exam_header, styles["Heading3"])
                    )
                    exam_table = Table(
                        [headers, row_values],
                        repeatRows=1
                    )
                    exam_table.setStyle(TableStyle([
                        ('GRID', (0,0), (-1,-1), 1, colors.black),
                        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
                        ('FONT', (0,0), (-1,0), 'Helvetica-Bold'),
                        ('ALIGN', (2,1), (-1,-1), 'CENTER'),
                        ('PADDING', (0,0), (-1,-1), 6),
                    ]))
                    elements.append(exam_table)
                    elements.append(Spacer(1, 14))
                    break
            r = header_row + 1
        else:
            r += 1
    if not found_student:
        print("❌ Student not found in any exam")
        return
    # ---------- FINAL RESULT ----------
    final_percent = round((combined_total / combined_max) * 100, 2)
    final_grade = calculate_grade(final_percent)
    final_result = "FAIL" if final_fail or final_grade == "E" else "PASS"
    elements.append(Spacer(1, 16))
    elements.append(Paragraph("FINAL RESULT", styles["Heading2"]))
    final_table = Table(
        [
            ["Total Marks", combined_total],
            ["Maximum Marks", combined_max],
            ["Percentage", f"{final_percent}%"],
            ["Final Grade", final_grade],
            ["Result", final_result],
        ],
        colWidths=[200, 200]
    )
    final_table.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
        ('FONT', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BACKGROUND', (0,-1), (-1,-1),
            colors.lightgreen if final_result == "PASS"
            else colors.salmon),
        ('ALIGN', (1,0), (-1,-1), 'CENTER'),
        ('PADDING', (0,0), (-1,-1), 8),
    ]))
    elements.append(final_table)
    doc.build(elements)
    print(f"✅ Final Report Card generated: {pdf_path}")

# managing function that manages marks and report card
def marks_management():
    while True:
        clear_screen()
        print("\n===== MARKS & EXAM MANAGEMENT =====")
        print("1. Create Exam")
        print("2. Enter Marks")
        print("3. Update Marks")
        print("4. Report  Generation")
        print("5. Back")
        choice = input("Enter choice (1-4): ")
        if choice == '1':
            create_exam()
        elif choice == '2':
            enter_marks()
        elif choice == '3':
            update_marks()
        elif choice == '4':
            generate_report_card_pdf()
        elif choice == '5':
            break
        else:
            print("❌ Invalid choice")

# teachers module that helps teacher to select what operation that he wants to perform
def teachermodule(teacher_name):
    while True:
        print(f"\n===== TEACHER MODULE ({teacher_name}) =====")
        print("1. View Assigned Classes")
        print("2. Update Profile Information")
        print("3. Manage Student Marks and Exams")
        print("4. Manage Attendance")
        print("5. Logout")
        choice = input("Enter choice (1-5): ")
        if choice == '1':
            clear_screen()
            view_assigned_classes(teacher_name)
        elif choice == '2':
            clear_screen()
            update_teacher_profile(teacher_name)
        elif choice == '3':
            clear_screen()
            marks_management()
        elif choice == '4':
            clear_screen()
            attendance_management()
        elif choice == '5':
            clear_screen()
            print("👋 Logged out successfully")
            break
        else:
            print("❌ Invalid choice")

# this ia admin panel start to select in which admin want to perform operation
#  in student class and teachers
def admin():
    clear_screen()
    print("Select In which you want to perform operation")
    print("1. Student Management")
    print("2. Teacher Management")
    print("3. Class Management")
    choice = (input("Enter your choice (1 or 2 or 3): "))
    if choice == '1':
        stdntmgmnt()
    elif choice == '2':
        teachermgmnt()
    elif choice=='3':
        classmgmnt()
    else:
        print("Invalid choice. Please select 1 or 2.")

# main module login module
# admin name and id is fix and teachers are check from teachers detail
# if they are present in excel sheet then they can enter otherwise not
def main():
    clear_screen()
    print("Welcome to School Management System ")
    username = input("Enter your username: ").strip()
    password = input("Enter your password: ").strip()
    if username == "admin" and password == "admin123":
        admin()
    else:
        teacher_name = teacher_login(username, password)
        if teacher_name:
            print(f"✅ Welcome {teacher_name}")
            teachermodule(teacher_name)
        else:
            print("❌ Invalid credentials. Access denied.")
main()