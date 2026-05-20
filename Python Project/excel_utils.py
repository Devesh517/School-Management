"""
excel_utils.py
Mirrors the original CLI project's folder/file layout exactly:

  excel_exports/
    Class_Management/
      class_master.xlsx
      Class_<name>/
        subjects.xlsx
        section_<X>.xlsx
    Teacher_Management/
      teacher_details.xlsx
      teacher_salary.xlsx
      teacher_assignments.xlsx
      removed_teachers.xlsx
    Attendance/
      Class_<name>/
        section_<X>.xlsx          (columns: Student_ID, Name, YYYY-MM-DD …)
    Marks/
      Class_<name>/
        section_<X>.xlsx
    Report_Cards/
      Class_<name>/Section_<X>/
        ReportCard_<class><sec>_<sid>.pdf
    Salary_Slips/
      SalarySlip_<name>_<month>.pdf
"""

import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from reportlab.platypus import Spacer, Table, TableStyle

BASE = "excel_exports"


# ─── helpers ───────────────────────────────────────────────────────────────

def _style(ws):
    """Bold header row, auto-width, wrap text."""
    bold = Font(bold=True)
    align = Alignment(vertical="center", wrap_text=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = align
            if cell.row == 1:
                cell.font = bold
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 4


def _wb(headers):
    """Create a new workbook with a styled header row."""
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    _style(ws)
    return wb, ws


def _ensure(path, headers):
    """Open existing workbook or create a new one with headers."""
    if os.path.exists(path):
        return load_workbook(path)
    wb, _ = _wb(headers)
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    return wb


# ─── Class Management ──────────────────────────────────────────────────────

def sync_class_master(classes):
    """
    Overwrite Class_Management/class_master.xlsx from DB rows.
    classes = [{"class_name":…, "total_sections":…, "status":…}, …]
    """
    path = os.path.join(BASE, "Class_Management", "class_master.xlsx")
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb, ws = _wb(["Class_Name", "Total_Sections", "Status"])
    for c in classes:
        ws.append([c["class_name"], c["total_sections"], c["status"]])
    _style(ws)
    wb.save(path)


def sync_subjects(class_name, subjects):
    """subjects = [{"subject_name":…}, …]"""
    folder = os.path.join(BASE, "Class_Management", f"Class_{class_name}")
    os.makedirs(folder, exist_ok=True)
    path = os.path.join(folder, "subjects.xlsx")
    wb, ws = _wb(["Subject_ID", "Subject_Name"])
    for i, s in enumerate(subjects, 1):
        ws.append([i, s["subject_name"]])
    _style(ws)
    wb.save(path)


def sync_section_students(class_name, section_name, students):
    """
    students = [{"student_id":…, "name":…, "dob":…, "age":…,
                  "mother_name":…, "father_name":…, "address":…,
                  "phone":…, "aadhar":…}, …]
    """
    folder = os.path.join(BASE, "Class_Management", f"Class_{class_name}")
    os.makedirs(folder, exist_ok=True)
    path = os.path.join(folder, f"section_{section_name}.xlsx")
    wb, ws = _wb([
        "Student_ID", "Name", "DOB", "Age",
        "Mother's Name", "Father's Name", "Address", "Phone No", "Aadhar NO"
    ])
    for s in students:
        ws.append([
            s.get("student_id"), s.get("name"), str(s.get("dob", "")),
            s.get("age"), s.get("mother_name"), s.get("father_name"),
            s.get("address"), s.get("phone"), s.get("aadhar"),
        ])
    _style(ws)
    wb.save(path)


# ─── Teacher Management ────────────────────────────────────────────────────

def sync_teacher_details(teachers):
    """teachers = list of teacher dicts from DB."""
    path = os.path.join(BASE, "Teacher_Management", "teacher_details.xlsx")
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb, ws = _wb([
        "ID", "Name", "Password", "DOB", "Age",
        "Phone", "Email", "Address", "Aadhar", "Account No", "Bank"
    ])
    for t in teachers:
        ws.append([
            t.get("teacher_id"), t.get("name"), t.get("password"),
            str(t.get("dob", "")), t.get("age"), t.get("phone"),
            t.get("email"), t.get("address"), t.get("aadhar"),
            t.get("account_number"), t.get("bank_name"),
        ])
    _style(ws)
    wb.save(path)


def sync_teacher_assignments(assignments):
    """assignments = list of assignment dicts."""
    path = os.path.join(BASE, "Teacher_Management", "teacher_assignments.xlsx")
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb, ws = _wb(["Teacher_ID", "Teacher_Name", "Class", "Section", "Subject", "Role"])
    for a in assignments:
        ws.append([
            a.get("teacher_id"), a.get("teacher_name"),
            a.get("class_name"), a.get("section_name"),
            a.get("subject_name"), a.get("role"),
        ])
    _style(ws)
    wb.save(path)


def sync_removed_teachers(removed):
    path = os.path.join(BASE, "Teacher_Management", "removed_teachers.xlsx")
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb, ws = _wb(["ID", "Name", "Reason", "Removed Date"])
    for r in removed:
        ws.append([r.get("teacher_id"), r.get("name"),
                   r.get("reason"), str(r.get("removed_date", ""))])
    _style(ws)
    wb.save(path)


def sync_salary(teacher_id, teacher_name, records):
    """records = list of salary dicts."""
    path = os.path.join(BASE, "Teacher_Management", "teacher_salary.xlsx")
    os.makedirs(os.path.dirname(path), exist_ok=True)
    # append-only: read existing, add new rows
    wb = _ensure(path, [
        "ID", "Name", "Month", "Basic", "Incentive", "Gross",
        "PF", "Professional_Tax", "TDS", "Total_Deduction", "Net_Salary"
    ])
    ws = wb.active
    existing_months = {
        (ws.cell(r, 1).value, ws.cell(r, 3).value)
        for r in range(2, ws.max_row + 1)
    }
    for rec in records:
        key = (rec.get("teacher_id"), rec.get("month"))
        if key not in existing_months:
            ws.append([
                rec.get("teacher_id"), rec.get("name"), rec.get("month"),
                rec.get("basic"), rec.get("incentive"), rec.get("gross"),
                rec.get("pf"), rec.get("professional_tax"), rec.get("tds"),
                rec.get("total_deduction"), rec.get("net_salary"),
            ])
    _style(ws)
    wb.save(path)


# ─── Attendance ────────────────────────────────────────────────────────────

def sync_attendance(class_name, section_name, date_str, records):
    """
    records = [{"student_id":…, "name":…, "status":"P"/"A"}, …]
    Adds a new date column to the section file (or updates existing).
    """
    folder = os.path.join(BASE, "Attendance", f"Class_{class_name}")
    os.makedirs(folder, exist_ok=True)
    path = os.path.join(folder, f"section_{section_name}.xlsx")

    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb, ws = _wb(["Student_ID", "Name"])
        # pre-fill students
        for r in records:
            ws.append([r["student_id"], r["name"]])

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    if date_str not in headers:
        date_col = ws.max_column + 1
        ws.cell(1, date_col).value = date_str
    else:
        date_col = headers.index(date_str) + 1

    # build lookup: student_id → row
    sid_row = {ws.cell(r, 1).value: r for r in range(2, ws.max_row + 1)}

    for rec in records:
        row_num = sid_row.get(rec["student_id"])
        if row_num:
            ws.cell(row_num, date_col).value = rec["status"]

    _style(ws)
    wb.save(path)


# ─── Marks ─────────────────────────────────────────────────────────────────

def sync_marks(class_name, section_name, exam_name, exam_date,
            max_marks, subjects, students_marks):

    """
    subjects = [{"subject_name":…}, …]

    students_marks = [
        {
            "student_id":…,
            "name":…,
            "marks":{"SubjectName": value, …},
            "total":…,
            "percentage":…,
            "grade":…
        }, …
    ]
    """

    folder = os.path.join(
        BASE,
        "Marks",
        f"Class_{class_name}"
    )

    os.makedirs(folder, exist_ok=True)

    path = os.path.join(
        folder,
        f"section_{section_name}.xlsx"
    )



    if os.path.exists(path):

        wb = load_workbook(path)

        ws = wb.active

    else:

        wb = Workbook()

        ws = wb.active



    # Exam Header

    header_text = (

        f"Exam: {exam_name} | "
        f"Exam Date: {exam_date} | "
        f"Max Marks: {max_marks}"

    )



    start_row = ws.max_row + 2 if ws.max_row > 1 else 1



    sub_names = [
        s["subject_name"]
        for s in subjects
    ]



    num_cols = len(sub_names) + 5



    ws.merge_cells(

        start_row=start_row,
        start_column=1,

        end_row=start_row,
        end_column=num_cols

    )



    ws.cell(start_row, 1).value = header_text



    # Table Headers

    col_headers = [

        "Student_ID",
        "Name"

    ] + sub_names + [

        "Total",
        "Percentage",
        "Grade"

    ]



    ws.append(col_headers)



    # Student Marks

    for sm in students_marks:

        row = [

            sm["student_id"],

            sm["name"]

        ]



        for sn in sub_names:

            row.append(

                sm["marks"].get(sn, "")

            )



        row += [

            sm.get("total", ""),

            sm.get("percentage", ""),

            sm.get("grade", "")

        ]



        ws.append(row)



    _style(ws)

    wb.save(path)


# ─── Salary Slip PDF ───────────────────────────────────────────────────────

def generate_salary_slip_pdf(record):
    """
    record = {teacher_id, name, month, basic, incentive, gross,
               pf, professional_tax, tds, total_deduction, net_salary}
    Returns the saved PDF path.
    """
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors

    folder = os.path.join(BASE, "Salary_Slips")
    os.makedirs(folder, exist_ok=True)
    name_safe = str(record["name"]).replace(" ", "_")
    month_safe = str(record["month"]).replace(" ", "_")
    pdf_path = os.path.join(folder, f"SalarySlip_{name_safe}_{month_safe}.pdf")

    doc = SimpleDocTemplate(pdf_path, rightMargin=36, leftMargin=36,
                            topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("<b>SALARY SLIP</b>", styles["Title"]))
    elements.append(Spacer(1, 15))

    info = Table([
        ["Teacher Name", record["name"]],
        ["Teacher ID",   record["teacher_id"]],
        ["Salary Month", record["month"]],
    ], colWidths=[150, 250])
    info.setStyle(TableStyle([
        ('GRID',       (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, 0), (0, -1),     colors.lightgrey),
        ('FONT',       (0, 0), (0, -1), 'Helvetica-Bold'),
        ('PADDING',    (0, 0), (-1, -1), 8),
    ]))
    elements.append(info)
    elements.append(Spacer(1, 20))

    salary = Table([
        ["Basic Salary",     f"₹{record['basic']}"],
        ["Incentive",        f"₹{record['incentive']}"],
        ["Gross Salary",     f"₹{record['gross']}"],
        ["PF (12%)",         f"₹{float(record['pf']):.2f}"],
        ["Professional Tax", f"₹{record['professional_tax']}"],
        ["TDS (5%)",         f"₹{float(record['tds']):.2f}"],
        ["Total Deduction",  f"₹{float(record['total_deduction']):.2f}"],
        ["NET SALARY",       f"₹{float(record['net_salary']):.2f}"],
    ], colWidths=[200, 200])
    salary.setStyle(TableStyle([
        ('GRID',       (0,  0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0,  0), (0,  -2),    colors.whitesmoke),
        ('BACKGROUND', (0, -1), (-1, -1),    colors.lightgrey),
        ('FONT',       (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('ALIGN',      (1,  0), (-1, -1), 'RIGHT'),
        ('PADDING',    (0,  0), (-1, -1), 8),
    ]))
    elements.append(salary)
    doc.build(elements)
    return pdf_path


# ─── Report Card PDF ───────────────────────────────────────────────────────

def generate_report_card_pdf(class_name, section_name, sid, student_name, exam_results):
    """
    exam_results = [
      {
        "exam_name": …, "start_date": …, "max_marks": …,
        "subjects": [{"subject_name":…, "marks_obtained":…}, …],
        "total": …, "max_total": …, "percentage": …, "grade": …
      }, …
    ]
    Returns saved PDF path.
    """
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib import colors

    folder = os.path.join(
        BASE, "Report_Cards", f"Class_{class_name}", f"Section_{section_name}"
    )
    os.makedirs(folder, exist_ok=True)
    pdf_path = os.path.join(
        folder, f"ReportCard_{class_name}{section_name}_{sid}.pdf"
    )

    doc = SimpleDocTemplate(pdf_path, rightMargin=36, leftMargin=36,
                            topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    center = ParagraphStyle("C", parent=styles["Title"], alignment=TA_CENTER)
    elements = []

    elements.append(Paragraph("REPORT CARD", center))
    elements.append(Spacer(1, 14))

    info = Table([
        ["Student Name", student_name],
        ["Class",        class_name],
        ["Section",      section_name],
        ["Student ID",   sid],
    ], colWidths=[120, 200])
    info.setStyle(TableStyle([
        ('GRID',       (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, 0), (0,  -1),    colors.lightgrey),
        ('FONT',       (0, 0), (0,  -1), 'Helvetica-Bold'),
        ('PADDING',    (0, 0), (-1, -1), 8),
    ]))
    elements.append(info)
    elements.append(Spacer(1, 20))

    combined_total = 0
    combined_max = 0
    final_fail = False

    for er in exam_results:
        combined_total += er["total"]
        combined_max   += er["max_total"]
        if er["grade"] == "E":
            final_fail = True

        elements.append(
            Paragraph(
                f"Exam: {er['exam_name']} | Date: {er['start_date']} | Max: {er['max_marks']}",
                styles["Heading3"]
            )
        )
        header = ["Subject"] + ["Marks"]
        rows = [[s["subject_name"], s["marks_obtained"]] for s in er["subjects"]]
        rows.append(["Total",      er["total"]])
        rows.append(["Percentage", f"{er['percentage']}%"])
        rows.append(["Grade",      er["grade"]])

        t = Table([header] + rows)
        t.setStyle(TableStyle([
            ('GRID',       (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, 0), (-1,  0),    colors.lightgrey),
            ('FONT',       (0, 0), (-1,  0), 'Helvetica-Bold'),
            ('PADDING',    (0, 0), (-1, -1), 6),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 14))

    # Final result
    final_pct = round((combined_total / combined_max) * 100, 2) if combined_max else 0
    final_grade = _calc_grade(final_pct)
    result = "FAIL" if final_fail or final_grade == "E" else "PASS"

    elements.append(Paragraph("FINAL RESULT", styles["Heading2"]))
    final_t = Table([
        ["Total Marks",   combined_total],
        ["Maximum Marks", combined_max],
        ["Percentage",    f"{final_pct}%"],
        ["Final Grade",   final_grade],
        ["Result",        result],
    ], colWidths=[200, 200])
    final_t.setStyle(TableStyle([
        ('GRID',       (0,  0), (-1, -1), 1, colors.black),
        ('FONT',       (0,  0), (-1,  0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1),
            colors.lightgreen if result == "PASS" else colors.salmon),
        ('ALIGN',      (1,  0), (-1, -1), 'RIGHT'),
        ('PADDING',    (0,  0), (-1, -1), 8),
    ]))
    elements.append(final_t)
    doc.build(elements)
    return pdf_path


def _calc_grade(pct):
    if pct >= 90: return "A1"
    if pct >= 80: return "A2"
    if pct >= 70: return "B1"
    if pct >= 60: return "B2"
    if pct >= 50: return "C"
    if pct >= 40: return "D"
    return "E"

def generate_exam_timetable_pdf(
        class_name,
        section_name,
        exam_name,
        rows
):

    from reportlab.platypus import (
        SimpleDocTemplate,
        Paragraph,
        Spacer,
        Table,
        TableStyle
    )

    from reportlab.lib.styles import getSampleStyleSheet

    from reportlab.lib import colors

    folder = os.path.join(BASE, "Exam_Timetables")

    os.makedirs(folder, exist_ok=True)

    path = os.path.join(
        folder,
        f"{exam_name}_{class_name}_{section_name}.pdf"
    )



    doc = SimpleDocTemplate(path)

    styles = getSampleStyleSheet()

    elements = []



    elements.append(
        Paragraph(
            f"<b>ABC SCHOOL</b><br/><br/>{exam_name} Examination Timetable",
            styles['Title']
        )
    )

    elements.append(Spacer(1, 20))



    elements.append(
        Paragraph(
            f"Class: {class_name}-{section_name}",
            styles['Heading2']
        )
    )

    elements.append(Spacer(1, 20))
    elements.append(
    Paragraph(
        "Exam Timing: 9:00 AM to 12:00 PM",
        styles['Heading3']
    )
)

    elements.append(Spacer(1, 10))


    data = [[
        "S.No",
    "Subject",
    "Date",
    "Start Time",
    "End Time"
    ]]



    for idx, r in enumerate(rows, start=1):

        data.append([

        idx,

        r["subject_name"],

        str(r["exam_date"]),

        str(r["start_time"]),

        str(r["end_time"])

])



    table = Table(data)



    table.setStyle(TableStyle([

        ('BACKGROUND', (0,0), (-1,0), colors.lightblue),

        ('TEXTCOLOR', (0,0), (-1,0), colors.black),

        ('GRID', (0,0), (-1,-1), 1, colors.black),

        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),

        ('BOTTOMPADDING', (0,0), (-1,0), 10),

    ]))



    elements.append(table)


    elements.append(Spacer(1, 40))

    signature_table = Table([[
    "Class Teacher Signature",
    "Principal Signature"
]])

    signature_table.setStyle(TableStyle([

    ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'),

    ('TOPPADDING', (0,0), (-1,-1), 30),

    ('ALIGN', (0,0), (-1,-1), 'CENTER')

]))

    elements.append(signature_table)
    doc.build(elements)



    return path